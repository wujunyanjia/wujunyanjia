"""
Project: 86python
Author:柠檬班-Tricy
Time:2021/12/1 19:50
Company:湖南零檬信息技术有限公司
Site: http://www.lemonban.com
Forum: http://testingpai.com 
"""
'''
接口的思路：-- 同步手工测试 
1、先有一个用例--接口测试用例参数 头部 url 等
2、从用例里读取数据 执行接口测试 
3、requests库执行接口测试 -- 执行结果
4、执行结果 预期结果 对比 ==  测试是否通过 -Pass  failed
5、最终的结果 --excel表格

接口自动化测试思路：--高级
1、Excel表格用例 --done
2、代码自动去Excel里读取出来  --done
3、传给 接口测试requests 方法去执行接口测试-结果  --函数--done
4、断言  -- 最终结果 
5、代码-结果自动回写到Excel --done 

封装函数步骤：
1、这个代码是否会重复使用  -- 封装，def 函数名（）
2、会变化的内容  --参数化
3、这个函数是否需要给别人数据传输  -- 定义返回值

    url_wuye ='http://47.115.15.198:7001/smarthome/user/register'  # 地址定义
    data_wuye = {"phone":"13656565757",
    "pwd":"1234567a",
    "rePwd":"1234567a",
    "userName":"柠檬86开4",
    "verificationCode":"lemon"}

从Excel里读取数据： 
openpyxl 库  ==操作Excel表格
安装  
导入 : 1）导入全部  2)导入部分 

excel表格两种 操作： 1） 读取数据  2）写入数据 
三个对象：
1、工作簿 -表格
2、sheet --表单
3、表格

写入数据： 写入后 进行数据保存  一定要关闭文件
wb = load_workbook("testcase_api_wuye.xlsx")  #加载工作簿-表本身,赋值给一个变量
sheet = wb["register"]  # 表单
cell = sheet.cell(row=1,column=1)  # 单元格
print(cell.value)
cell.value = "用例编号"  # 先得到这个单元格的内容  --对他进行重新赋值
wb.save("testcase_api_wuye.xlsx")  # 保存 文件  == 保存文件之前一定要关闭这个文件！

'''
import requests,pprint,openpyxl
# import openpyxl
from openpyxl import load_workbook   # 部分导入
openpyxl.load_workbook("testcase_api_wuye.xlsx")

# 发送接口请求函数封装
def api_fun(url_api,data_api):
    header_wy = {"X-Lemonban-Media-Type":"lemonban.v2","Content-Type":"application/json"}
    res = requests.post(url=url_api,json=data_api,headers=header_wy)
    return res.json()  # 把接口的执行结果定义为返回值

# 读取excel数据函数封装
def read_date(filename,sheetname):
    wb = load_workbook(filename)  #加载工作簿-表本身,赋值给一个变量
    sheet = wb[sheetname]  # 表单
    max_row = sheet.max_row  # 获取表单的里最大行数
    case_list = []  # 空列表 --装数据
    for i in range(2,max_row+1):
        dict1 = dict(
        case_id = sheet.cell(row=i,column=1).value,
        url = sheet.cell(row=i,column=5).value, # 单元格
        data = sheet.cell(row=i,column=6).value, # 单元格
        expe_result = sheet.cell(row=i,column=7).value) # 单元格
        case_list.append(dict1)
    return  case_list

# 结果回写excel函数
def write_result(filename,sheetname,row,column,new_value):
    wb = load_workbook(filename)
    sheet = wb[sheetname]
    sheet.cell(row,column).value  = new_value
    wb.save(filename)

'''
{'case_id': 1, 
'url': 'http://47.115.15.198:7001/smarthome/user/register', 
'data': '{"phone":"17750542236","pwd":"1234567a","rePwd":"1234567a","userName":"柠檬小可爱1","verificationCode":"611203"}', 
'expe_result': '{"code":"0","msg":"操作成功"}'}

eval()内置函数：""包裹的Python的表达式 --  运行出来  

{'code': '1008', 'msg': '注册失败，手机号码已被使用！', 'data': None}
'''
# 执行并断言的函数封装
def execute_fun(filename,sheetname):
    cases = read_date(filename,sheetname)  # 调用读取函数，并得到他的返回值
    for case in cases:
       case_id =  case.get("case_id")
       url = case["url"]
       data = case.get("data")  # 参数是从excel表里取 -- 文本-字符串 ==  转化为字典
       data = eval(data)  # 字符串 脱掉引号 -- 字典
       exp_re = case.get("expe_result")
       exp_re = eval(exp_re)  # 字符串 脱掉引号 -- 字典
       real_res = api_fun(url_api=url,data_api=data)  # 发送接口请求
       real_msg = real_res.get("msg")
       exp_msg = exp_re.get("msg")
       print("执行结果：{}".format(real_msg))
       print("预期结果：{}".format(exp_msg))
       if real_msg == exp_msg:  # 断言
           print("第{}测试用例执行通过！".format(case_id))
           real_result = "Passed"
       else:
           print("第{}条测试用例执行不通过！".format(case_id))
           real_result = "Failed"
       print("*" * 20)
       write_result(filename,sheetname,case_id+1,8,real_result)  # 调用回写结果的函数

execute_fun("testcase_api_wuye.xlsx","login")











